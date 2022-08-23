import openpyxl
import pandas as pd
import pyautogui
import os


#1PASSO - Execute o arquivo 'Validar Tabela.exe' para validar as tabelas salvas do sistema antes

def macro_validador():
    pyautogui.press('enter')
    pyautogui.hotkey('alt', 'f4')
    pyautogui.press('tab')
    pyautogui.press('enter')

    
def validacao_tabelas_xls():
    if os.path.isfile('Converter/produto.xls'):
        os.startfile('Converter\produto.xls')


#2°PASSO - Converter as tabelas validadas de .xls para .xlsx

def converter_produtos_preco():
    if os.path.isfile("Converter/produto.xls"):
        pl = pd.read_excel("Converter/produto.xls")
        cv = pd.ExcelWriter('Tratamento/produto.xlsx')
        pl.to_excel(cv, 'Produtos_Preço', index=False)
        cv.save()
    else:
        return

    
#3°PASSO - Criar tabelas escopo onde vamos armazenar os dados das tabelas convertidas
    

def criar_produtos_preco():
    if os.path.isfile('Converter/produto.xls'):
        new = openpyxl.Workbook()
        linha = new['Sheet']
        linha.append(['CODIGO', 'CODIGO_BARRAS', 'DESCRICAO', 'VALOR_1', 'VALOR_2', 'VALOR_3', 'VALOR_4', 'VALOR_5'])
        new.save('Escopo/corpo_produto.xlsx')


#4°PASSO - Fazer o tratamento dos dados, corrigindo e passando tudo para as tabelas escopo

def tratamento_produtos_preco():
    if os.path.isfile('Tratamento/produto.xlsx'):
        planilha_produto = openpyxl.load_workbook('Tratamento/produto.xlsx')
        new = openpyxl.load_workbook('Escopo/corpo_produto.xlsx')
        nova_linha = new['Sheet']
        linha_produto = planilha_produto['Produtos_Preço']

        for row in linha_produto.iter_rows():

            if row[16].value == None or row[16].value == "VALOR 01" or row[16].value == "VALOR 02" or row[16].value == "VALOR 03" or row[16].value == "VALOR 04" or row[16].value == "VALOR 05" or row[13].value == None or row[13].value == "VALOR 01" or row[13].value == "VALOR 02" or row[13].value == "VALOR 03" or row[13].value == "VALOR 04" or row[13].value == "VALOR 05" or row[11].value == None or  row[11].value == "VALOR 01" or row[11].value == "VALOR 02" or row[11].value == "VALOR 03" or row[11].value == "VALOR 04" or row[11].value == "VALOR 05"  or row[10].value == None or row[10].value == "VALOR 01" or row[10].value == "VALOR 02" or row[10].value == "VALOR 03" or row[10].value == "VALOR 04" or row[10].value == "VALOR 05" or row[9].value == None or row[9].value == "VALOR 01" or row[9].value == "VALOR 02" or row[9].value == "VALOR 03" or row[9].value == "VALOR 04" or row[9].value == "VALOR 05" or row[3].value == "DESCRIÇÃO" or row[1].value == "CÓDIGO BARRAS" or row[0].value == "Kyacom Informática" or row[0].value == None or row[0].value == "Seção:" or row[0].value == "CÓDIGO" or row[0].value == "  FORNECEDOR={0}     SITUAÇÃO={PRODUTOS ATIVOS}     ORDEM={CÓDIGO DO PRODUTO}" or row[0].value == "RELATÓRIO DE TABELA DE PREÇOS - (REL.10)" or row[0].value == "KYACOM":

                row[0].value == ''
                row[1].value == ''
                row[3].value == ''
                row[9].value == ''
                row[10].value == ''
                row[11].value == ''
                row[13].value == ''
                row[16].value == ''

            else:

                c1 = row[9].value.replace(".","").replace(",",".")
                V1 = float(c1)
                c2 = row[10].value.replace(".","").replace(",",".")
                V2 = float(c2)
                c3 = row[11].value.replace(".","").replace(",",".")
                V3 = float(c3)
                c4 = row[13].value.replace(".","").replace(",",".")
                V4 = float(c4)
                c5 = row[16].value.replace(".","").replace(",",".")
                V5 = float(c5)

                nova_linha.append([row[0].value, row[1].value,row[3].value, V1,V2,V3,V4,V5])

        new.save('Relatorio/Produto.csv')


#Chamando as funcoes de cada passo a passo

#1°PASSO--------------------
validacao_tabelas_xls()
macro_validador()

#2°PASSO--------------------    
converter_produtos_preco()

#3°PASSO--------------------
criar_produtos_preco()

#4°PASSO--------------------
tratamento_produtos_preco()


