import openpyxl
import pandas as pd
import os
#1PASSO - Execute o arquivo 'Validar Tabela.exe' para validar as tabelas salvas do sistema antes

#2°PASSO - Converter as tabelas validadas de .xls para .xlsx

def converter_inventario_falta():
    if os.path.isfile("Converter/inventario-falta.xls"):
        pl = pd.read_excel("Converter/inventario-falta.xls")
        cv = pd.ExcelWriter('Tratamento/inventario-falta.xlsx')
        pl.to_excel(cv, 'Inventario', index=False)
        cv.save()
    else:
        return

def converter_produtos_preco():
    if os.path.isfile("Converter/produto-preco.xls"):
        pl = pd.read_excel("Converter/produto-preco.xls")
        cv = pd.ExcelWriter('Tratamento/produto-preco.xlsx')
        pl.to_excel(cv, 'Produtos_Preço', index=False)
        cv.save()
    else:
        return

    
#3°PASSO - Criar tabelas escopo onde vamos armazenar os dados das tabelas convertidas
    
def criar_inventario_faltas():
    if os.path.isfile('Converter/inventario-falta.xls'):
        new = openpyxl.Workbook()
        linha = new['Sheet']
        linha.append(['CODIGO_BARRAS', 'DESCRICAO', 'ESTOQUE', 'CONTAGEM', 'FALTA'])
        new.save('DataFrames/Estruturas/corpo_inventario_falta.xlsx')
        
def criar_produtos_preco():
    if os.path.isfile('Converter/produto-preco.xls'):
        new = openpyxl.Workbook()
        linha = new['Sheet']
        linha.append(['CODIGO', 'CODIGO_BARRAS', 'DESCRICAO', 'VALOR_1', 'VALOR_2', 'VALOR_3', 'VALOR_4', 'VALOR_5'])
        new.save('DataFrames/Estruturas/corpo_produto-preco.xlsx')


#4°PASSO - Fazer o tratamento dos dados, corrigindo e passando tudo para as tabelas escopo

def converter_produtos_preco():
    if os.path.isfile('Tratamento/produto-preco.xlsx'):
        planilha_produto = openpyxl.load_workbook('Tratamento/produto-preco.xlsx')
        new = openpyxl.load_workbook('DataFrames/Estruturas/corpo_produto-preco.xlsx')
        nova_linha = new['Sheet']
        linha_produto = planilha_produto['Produtos_Preço']

        for row in linha_produto.iter_rows():

            if row[16].value == None or row[16].value == "VALOR 05" or row[13].value == None or row[13].value == "VALOR 04" or row[11].value == None or row[11].value == "VALOR 03" or row[10].value == None or row[10].value == "VALOR 02" or row[9].value == None or row[9].value == "VALOR 01" or row[3].value == "DESCRIÇÃO" or row[1].value == "CÓDIGO BARRAS" or row[0].value == "Kyacom Informática" or row[0].value == None or row[0].value == "Seção:" or row[0].value == "CÓDIGO" or row[0].value == "  FORNECEDOR={0}     SITUAÇÃO={PRODUTOS ATIVOS}     ORDEM={CÓDIGO DO PRODUTO}" or row[0].value == "RELATÓRIO DE TABELA DE PREÇOS - (REL.10)" or row[0].value == "KYACOM":

                row[0].value == ''
                row[1].value == ''
                row[3].value == ''
                row[9].value == ''
                row[10].value == ''
                row[11].value == ''
                row[13].value == ''
                row[16].value == ''

            else:

                c1 = row[9].value.replace(".","").replace(",",".")
                V1 = float(c1)
                c2 = row[10].value.replace(".","").replace(",",".")
                V2 = float(c2)
                c3 = row[11].value.replace(".","").replace(",",".")
                V3 = float(c3)
                c4 = row[13].value.replace(".","").replace(",",".")
                V4 = float(c4)
                c5 = row[16].value.replace(".","").replace(",",".")
                V5 = float(c5)

                nova_linha.append([row[0].value, row[1].value,row[3].value, V1,V2,V3,V4,V5])

        new.save('DataFrames/Bases/Produto-Preco.xlsx')
    else:
        return

def converter_inventario_falta():
    if os.path.isfile('Tratamento/inventario-falta.xlsx'):
        planilha_inventario = openpyxl.load_workbook('Tratamento/inventario-falta.xlsx')
        new = openpyxl.load_workbook('DataFrames/Estruturas/corpo_inventario_falta.xlsx')
        nova_linha = new['Sheet']
        linha_inventario = planilha_inventario['Inventario']

        for row in linha_inventario.iter_rows():
            if row[10].value == None or row[7].value == None or row[3].value == None or row[10].value == "FALTA" or row[7].value == "CONTAGEM" or row[3].value == "ESTOQUE" or row[1].value == "DESCRICAO" or row[0].value == "KYACOM" or row[1].value == "" or row[1].value == None or row[0].value == "CODIGO BARRAS" or row[0].value == "Kyacom Informática":
                row[0].value == ''
                row[1].value == ''
                row[3].value == ''
                row[7].value == ''
                row[10].value == ''
            
            elif row[0].value == '' or row[0].value == None or row[0].value == 'Nan':
                row[0].value == ''
                
            else:

                c1 = row[3].value.replace(".","").replace(",",".")
                E = float(c1)
                c2 = row[7].value.replace(".","").replace(",",".")
                C = float(c2)
                c3 = row[10].value.replace(".","").replace(",",".")
                F = float(c3)

                nova_linha.append([row[0].value, row[1].value, E, C, F])

        new.save('DataFrames/Bases/Inventario-Falta.xlsx')
    else:
        return
    
#5°PASSO - Mesclar as tabelas escopo para gerar uma nova tabela , ja fazendo toda a analise de dados para as comparacoes e os calculos armazenados em novas colunas
    
def mesclar_tabelas():
    tab1 = pd.read_excel('DataFrames/Bases/Produto-Preco.xlsx')
    tab2 = pd.read_excel('DataFrames/Bases/Inventario-Falta.xlsx')

    new = tab1.merge(tab2, left_on=['CODIGO_BARRAS','DESCRICAO'], right_on=['CODIGO_BARRAS','DESCRICAO'],how='outer')

    new['VALOR_1_FALTA'] = new['FALTA'] * new['VALOR_1']
    new['VALOR_2_FALTA'] = new['FALTA'] * new['VALOR_2']
    new['VALOR_3_FALTA'] = new['FALTA'] * new['VALOR_3']
    new['VALOR_4_FALTA'] = new['FALTA'] * new['VALOR_4']
    new['VALOR_5_FALTA'] = new['FALTA'] * new['VALOR_5']


    new.to_excel('Relatorios_Convertidos/Relatorio.xlsx')
    
#6°PASSO - Fazer correcoes da nova tabela e nesse caso acrescentar campo TOTAL com os valores totais de determinadas colunas da tabela

def correcao_tabela():
    pla = openpyxl.load_workbook('Relatorios_Convertidos/Relatorio.xlsx')

    linhas = pla['Sheet1']
    v1 = v2 = v3 = v4 = v5 = 0
    v1f = v2f = v3f = v4f = v5f = 0

    for row in linhas.iter_rows():

        if row[16].value == 'VALOR_5_FALTA' or row[15].value == 'VALOR_4_FALTA' or row[14].value == 'VALOR_3_FALTA' or row[13].value == 'VALOR_2_FALTA' or row[12].value == 'VALOR_1_FALTA' or row[4].value == 'VALOR_1' or row[5].value == 'VALOR_2' or row[6].value == 'VALOR_3' or row[7].value == 'VALOR_4' or row[5].value == 'VALOR_5':
            print('Correcao')
        
        elif row[16].value == None or row[15].value == None or row[14].value == None or row[13].value == None or row[12].value == None or row[4].value == None  or row[5].value == None or row[6].value == None or row[7].value == None or row[5].value == None:
            row[4].value = 0
            row[5].value = 0
            row[6].value = 0
            row[7].value = 0
            row[8].value = 0
            row[12].value = 0
            row[13].value = 0
            row[14].value = 0
            row[15].value = 0
            row[16].value = 0

        else:
            v1 = v1 + row[4].value
            v2 = v2 + row[5].value
            v3 = v3 + row[6].value
            v4 = v4 + row[7].value
            v5 = v5 + row[8].value
            v1f = v2f + row[12].value
            v2f = v2f + row[13].value
            v3f = v3f + row[14].value
            v4f = v4f + row[15].value
            v5f = v5f + row[16].value


    linhas.append(['TOTAL', '', '', '', v1, v2 , v3, v4, v5, '', '', '', v1f, v2f, v3f, v4f, v5f])
    pla.save('Relatorios_Convertidos/Relatorio.xlsx')


#Chamando as funcoes de cada passo a passo

#2°PASSO--------------------    
converter_produtos_preco()
converter_inventario_falta()

#3°PASSO--------------------
criar_inventario_faltas()
criar_produtos_preco()

#4°PASSO--------------------
converter_inventario_falta()
converter_produtos_preco()

#5°PASSO--------------------
mesclar_tabelas()

#6°PASSO--------------------
correcao_tabela()
